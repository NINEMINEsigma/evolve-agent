"""Excel 工具 — 读写 .xlsx 工作簿。

模块导入时通过 ``registry.register()`` 注册。
依赖 ``openpyxl``（请参见 requirements.txt）。
"""

from __future__ import annotations

import logging
from typing import Any, Dict, List

from abstract.tools.registry import registry, tool_error, tool_result
from component.tools.filesystem import _s as _get_sandbox

logger = logging.getLogger(__name__)

# 延迟导入，启动时若未安装则以友好错误提示替代崩溃
_openpyxl = None
try:
    import openpyxl as _openpyxl
except ImportError:
    pass


# ---------------------------------------------------------------------------
# Handlers
# ---------------------------------------------------------------------------


def _handle_read_excel(args: dict[str, Any]) -> dict:
    path: str = str(args.get("path", "")).strip()
    if not path:
        return tool_error("path is required")

    if _openpyxl is None:
        return tool_error(
            "openpyxl is not installed. Run: pip install openpyxl",
        )

    try:
        sandbox = _get_sandbox()
        r = sandbox.resolve_read(path)
    except Exception as e:
        return tool_error(str(e), path=path)

    try:
        wb = _openpyxl.load_workbook(r.real, read_only=True, data_only=True)
    except Exception as e:
        return tool_error(f"Failed to open workbook: {e}", path=path)

    try:
        sheets: dict[str, list[dict[str, Any]]] = {}
        sheet_names: list[str] = wb.sheetnames

        selected: str = (args.get("sheet") or "").strip()
        target_sheets: list[str] = [selected] if selected else sheet_names

        for name in target_sheets:
            if name not in wb:
                return tool_error(f"Sheet '{name}' not found. Available: {sheet_names}")

            ws = wb[name]
            rows_iter = ws.iter_rows(values_only=True)
            headers: list[str] | None = None
            sheet_rows: list[dict[str, Any]] = []

            for row_idx, row in enumerate(rows_iter):
                if row_idx == 0:
                    # 首行作为表头
                    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(row)]
                    continue
                if all(v is None for v in row):
                    continue  # 跳过空行
                # headers 可能为 None，需要类型断言
                sheet_rows.append(
                    {headers[i]: (v if v is not None else None) for i, v in enumerate(row)}  # type: ignore
                )

            sheets[name] = sheet_rows

        return tool_result(
            path=path,
            sheets=sheets,
            sheet_names=sheet_names,
            active_sheet=wb.active.title,
        )
    finally:
        wb.close()


def _handle_write_excel(args: dict[str, Any]) -> dict:
    path: str = str(args.get("path", "")).strip()
    data: list[dict[str, Any]] | dict[str, list[dict[str, Any]]] = args.get("data", [])

    if not path:
        return tool_error("path is required")

    if _openpyxl is None:
        return tool_error(
            "openpyxl is not installed. Run: pip install openpyxl",
        )

    try:
        sandbox = _get_sandbox()
        r = sandbox.resolve_write(path)
        r.real.parent.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return tool_error(str(e), path=path)

    wb = _openpyxl.Workbook()
    # 默认 sheet 会在首次创建时被覆盖
    wb.remove(wb.active)

    try:
        # 支持单 sheet 和多 sheet 两种输入形态
        sheets_input: dict[str, list[dict[str, Any]]]
        if isinstance(data, dict):
            sheets_input = data
        elif isinstance(data, list) and data:
            # 单 sheet — 以 data 作为 sheet 名
            sheet_name: str = str(args.get("sheet", "Sheet1"))
            sheets_input = {sheet_name: data}
        else:
            return tool_error(
                "data must be a non-empty list of objects, "
                "or a dict of {sheet_name: rows}"
            )

        for sheet_name, rows in sheets_input.items():
            if not rows or not isinstance(rows, list):
                continue
            ws = wb.create_sheet(title=sheet_name)

            columns: list[str] = args.get("columns") or list(rows[0].keys())
            # 表头
            ws.append(columns)
            # 数据行
            for row in rows:
                ws.append([row.get(col, "") for col in columns])

        wb.save(str(r.real))
    except Exception as e:
        return tool_error(f"Failed to write workbook: {e}", path=path)
    finally:
        wb.close()

    return tool_result(
        path=path,
        sheets=list(sheets_input.keys()) if isinstance(data, dict) else [args.get("sheet", "Sheet1")],
        success=True,
    )


# ---------------------------------------------------------------------------
# Registration
# ---------------------------------------------------------------------------

_EXCEL_PARAMS: dict = {
    "type": "object",
    "properties": {
        "path": {
            "type": "string",
            # Excel 文件逻辑路径，必须使用命名空间前缀（ws:、fork:）。例如 'ws:data/report.xlsx'。
            "description": """Excel file logical path, must use a namespace prefix (ws:, fork:). E.g. 'ws:data/report.xlsx'.""",
        },
    },
    "required": ["path"],
}

registry.register(
    name="read_excel",
    toolset="extools",
    schema={
        # 读取 .xlsx Excel 工作簿。
        #
        # ## 前置条件
        # 必须安装 openpyxl。
        # path 必须使用命名空间前缀（如 ws:、fork:）。
        #
        # ## 调用效果
        # 读取工作簿，首行作为列名，返回所有 sheet 或指定 sheet 的数据。
        # 空行会被跳过。使用 data_only=True 读取公式结果而非公式本身。
        #
        # ## 返回
        # ```json
        # {"path": "ws:data/report.xlsx", "sheets": {"Sheet1": [{"col1": "value1"}]}, "sheet_names": ["Sheet1"], "active_sheet": "Sheet1"}
        # ```
        #
        # ## 何时使用
        # - 从 Excel 文件读取结构化数据。
        # - 批量处理表格内容。
        #
        # ## 副作用/注意
        # - 只读操作，不会修改源文件。
        # - 公式返回计算后的值。
        "description": """Read a .xlsx Excel workbook.

## Prerequisites
openpyxl must be installed. The path must use a namespace prefix (e.g. ws:, fork:).

## Effect
Reads the workbook and returns data from all sheets or a single specified sheet. The first row is used as column names. Empty rows are skipped. Uses data_only=True to read formula results instead of formulas.

## Returns
```json
{"path": "ws:data/report.xlsx", "sheets": {"Sheet1": [{"col1": "value1"}]}, "sheet_names": ["Sheet1"], "active_sheet": "Sheet1"}
```

## When to Use
- Read structured data from an Excel file.
- Batch process tabular content.

## Side Effects / Notes
- Read-only operation; does not modify the source file.
- Formulas return their calculated values.""",
        "parameters": {
            **_EXCEL_PARAMS,
            "properties": {
                **_EXCEL_PARAMS["properties"],
                "sheet": {
                    "type": "string",
                    # 要读取的 sheet 名称。省略时返回所有 sheet。
                    "description": """Name of the sheet to read. Omit to return all sheets.""",
                },
            },
        },
    },
    handler=_handle_read_excel,
    emoji="📗",
)

registry.register(
    name="write_excel",
    toolset="extools",
    schema={
        # 将结构化数据写入 .xlsx Excel 文件。
        #
        # ## 前置条件
        # 必须安装 openpyxl。
        # path 必须使用命名空间前缀（如 ws:、fork:）。
        #
        # ## 调用效果
        # 支持两种输入格式：
        # 1) 单 sheet：data 为行对象列表，默认 sheet 名 'Sheet1'。
        # 2) 多 sheet：data 为 {sheet_name: rows} 字典。
        # 每行是一个字典，第一行自动作为表头写入。columns 可控制列顺序和白名单。
        #
        # ## 返回
        # ```json
        # {"path": "ws:data/report.xlsx", "sheets": ["Sheet1"], "success": true}
        # ```
        #
        # ## 何时使用
        # - 将表格数据导出为 Excel。
        # - 生成多 sheet 报表。
        #
        # ## 副作用/注意
        # - 会写入新文件，可能覆盖同名文件。
        # - 只有 columns 中列出的字段会被写入。
        "description": """Write structured data to a .xlsx Excel file.

## Prerequisites
openpyxl must be installed. The path must use a namespace prefix (e.g. ws:, fork:).

## Effect
Supports two input formats:
1) Single sheet: data is a list of row objects, default sheet name 'Sheet1'.
2) Multi sheet: data is a {sheet_name: rows} dict.
Each row is a dict; the first row's keys become headers automatically. columns controls column order and allowlist.

## Returns
```json
{"path": "ws:data/report.xlsx", "sheets": ["Sheet1"], "success": true}
```

## When to Use
- Export tabular data to Excel.
- Generate multi-sheet reports.

## Side Effects / Notes
- Writes a new file and may overwrite an existing file with the same name.
- Only fields listed in columns are written.""",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    # Excel 文件逻辑路径（ws: 或 fork: 前缀）。
                    "description": """Excel file logical path (ws: or fork: prefix).""",
                },
                "data": {
                    "type": "array",
                    "items": {"type": "object"},
                    # 要写入的数据。可以是行对象列表（单 sheet）或 {sheet_name: rows} 字典（多 sheet）。
                    "description": """Data to write. Can be a list of row objects (single sheet) or a {sheet_name: rows} dict (multi sheet).""",
                },
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    # 列顺序及白名单。省略时使用第一行字典的键顺序。
                    "description": """Column order and allowlist. When omitted, uses keys from the first row dict.""",
                },
                "sheet": {
                    "type": "string",
                    # 单 sheet 模式下的 sheet 名称（默认 'Sheet1'）。当 data 为字典时忽略此参数。
                    "description": """Sheet name for single sheet mode (default 'Sheet1'). Ignored when data is a dict.""",
                },
            },
            "required": ["path", "data"],
        },
    },
    handler=_handle_write_excel,
    emoji="📗",
    danger_level="write",
)